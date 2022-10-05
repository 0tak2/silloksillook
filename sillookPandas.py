import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from sillookArticleEntity import SillookArticleEntity
from sillookParser import makeUrlFromArticleId

def entitiyToList(item: SillookArticleEntity):
    dbID = str(item.getId())
    location = item.getLocation()
    content_kor = item.getContentKor()
    content_han = item.getContentHan()
    note = item.getNote()
    metadata = item.getMetadata()
    articleId = item.getArticleId()
    url = makeUrlFromArticleId(articleId)
    title = item.getTitle()
    url_cell_content = '=HYPERLINK("{}", "{}")'.format(url, articleId)

    return [dbID, title, location, content_kor,
        content_han, note, metadata, url_cell_content]

def entitiesTo2DList(entitiesList):
    result = list(map(entitiyToList, entitiesList))
    return result

def makeDf(TwoDList):
    df = pd.DataFrame(TwoDList, columns =['순번', '기사제목', '기사위치', '국문내용', '원문내용', '메모', '메타데이터', '원본 링크'])
    return df

def dfToExcel(df, fileName, sheet_name):
    df.to_excel(fileName, sheet_name=sheet_name, index=False)

    # 스타일링을 위해 다시 파일을 연다
    wb = load_workbook(fileName)
    ws = wb.active

    # 각 열에 최적의 너비를 설정한다
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 23
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 9

    # 모든 셀에 자동 줄바꿈, 세로 정렬을 설정한다
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

   # 헤더에 가운데 정렬을 설정한다
    for cell in ws[1]:
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    # A열에 가운데 정렬을 설정한다
    for row in ws.iter_rows():
        row[0].alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    wb.save(fileName)
